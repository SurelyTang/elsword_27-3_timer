import tkinter as tk
import keyboard 
import pyttsx3
import threading 
from openpyxl import load_workbook
import winsound

font_size = 10 
file_path = '计时器配置表.xlsx' 
vertical_flag = 0
workbook = load_workbook(file_path)
sheet = workbook.worksheets[0] 
font_size_flag = sheet.cell(2,7).value
button_flag = sheet.cell(2,8).value
vertical_flag = sheet.cell(2,9).value
if (font_size_flag!=0):
    font_size = font_size_flag

num_to_symbol = {
    '1': '!',
    '2': '@',
    '3': '#',
    '4': '$',
    '5': '%',
    '6': '^',
    '7': '&',
    '8': '*',
    '9': '(',
    '0': ')',
    "NUM0": 82,
    "NUM1": 79,
    "NUM3": 81,
    "NUM5": 76,
    "NUM7": 71,
    "NUM9": 73,
    "+":78,
}
numeric_string = "1234567890"

def close_window():
    root.destroy()

def increase_font_size():
    global font_size
    font_size += 2  # Increase the font size by 2
    for timer in timers:
        timer.update_font_size()
        

def decrease_font_size():
    global font_size
    font_size -= 2  # Increase the font size by 2
    if font_size <= 5 :
        font_size = 5
    for timer in timers:
        timer.update_font_size()

class CountdownTimer:
    def __init__(self, root, timer_name, initial_seconds, row, column, hotkeys, flag_20, flag_10, flag_0):
        self.root = root
        self.timer_name = timer_name
        self.remaining_time = tk.StringVar()
        self.initial_seconds = initial_seconds
        self.remaining_seconds = initial_seconds
        if (vertical_flag):
            temp = row
            row = column
            column = temp
        self.time_label = tk.Label(root, textvariable=self.remaining_time, font=("Helvetica", font_size), width=12)
        self.time_label.grid(row=row, column=column, padx=5, pady=5)
        
        if (button_flag == 1):
            self.start_button = tk.Button(root, text=f"重置计时 {timer_name}", command=self.reset_and_start, font=("Times", font_size))
            self.start_button.grid(row=row+(1+vertical_flag)%2, column=column+vertical_flag, padx=5, pady=5)

        self.is_running = False
        self.timer_id = None
        self.engine = pyttsx3.init()
        self.hotkeys = hotkeys 
        if hotkeys:  
            self.bind_hotkey(hotkeys)
        self.flag_20 = flag_20
        self.flag_10 = flag_10
        self.flag_0 = flag_0
        self.loop = 0
        if timer_name == "超越栏":
            self.loop = 1

    
    def reset_and_start(self):
        self.reset_countdown()
        self.start_countdown()
        
        
    def start_countdown(self):
        if not self.is_running:
            self.is_running = True
            if self.timer_id is not None:
                self.root.after_cancel(self.timer_id)  # 取消先前的定时器
            self.update_clock()

    def reset_countdown(self):
        self.is_running = False
        self.remaining_seconds = self.initial_seconds
        mins, secs = divmod(self.remaining_seconds, 60)
        timeformat = '{:02d}:{:02d}'.format(mins, secs)
        self.remaining_time.set(f"{self.timer_name}: {timeformat}")

    def update_clock(self):
        if self.remaining_seconds >= 0 and self.is_running:
            mins, secs = divmod(self.remaining_seconds, 60)
            timeformat = '{:02d}:{:02d}'.format(mins, secs)
            self.remaining_time.set(f"{self.timer_name}: {timeformat}")
            
            if self.remaining_seconds <= 20:
                if self.remaining_seconds == 20 and self.flag_20 == 1:
                        tts_thread = threading.Thread(target=self.speak_timer_name_20)
                        tts_thread.start()
            
            if self.remaining_seconds <= 15:
                self.time_label.config(fg="red")  # 改成红色
                if self.remaining_seconds == 10 and self.flag_10 == 1:
                    tts_thread = threading.Thread(target=self.speak_timer_name_10)
                    tts_thread.start()
                
                if self.loop == 1:
                    if self.remaining_seconds == 3:
                        tts_thread = threading.Thread(target=self.speak_timer_name_3)
                        tts_thread.start()
                
                if self.remaining_seconds == 0:
                    if self.loop == 1:
                        self.remaining_seconds = self.initial_seconds
                    elif self.flag_0 == 1:
                        tts_thread = threading.Thread(target=self.speak_timer_name)
                        tts_thread.start()
                    
            else:
                self.time_label.config(fg="black")  # 字体黑色

            self.remaining_seconds -= 1
            self.timer_id = self.root.after(1000, self.update_clock)  # 更新计时器

        elif not self.is_running:
            self.is_running = False
            
    def update_font_size(self):
        self.time_label.config(font=("Helvetica", font_size))
        self.start_button.config(font=("Times", font_size))
    
    def speak_timer_name_20(self):
        # Initialize TTS engine
        engine = pyttsx3.init()
        engine.say(self.timer_name + "二十秒")
        engine.runAndWait()
            
    def speak_timer_name_10(self):
        # Initialize TTS engine
        engine = pyttsx3.init()
        engine.say(self.timer_name + "十秒")
        engine.runAndWait()
    
    def speak_timer_name_5(self):
        # Initialize TTS engine
        engine = pyttsx3.init()
        engine.say(self.timer_name + "五秒")
        engine.runAndWait()
        
    def speak_timer_name_3(self):
        winsound.PlaySound('10880.wav',winsound.SND_FILENAME)
        
    def speak_timer_name(self):
        # Initialize TTS engine
        engine = pyttsx3.init()
        engine.say(self.timer_name)
        engine.runAndWait()
    
    def on_press(self):
        self.reset_countdown()
        self.start_countdown()
    
    def bind_hotkey(self, hotkeys):
        hotkeys = self.hotkeys
        for hotkey in hotkeys:
            keyboard.add_hotkey(hotkey, self.on_press)


def on_test_press(event):
    for timer in timers:
        if timer.is_running == True:
            timer.is_running = False
    


if __name__ == '__main__':
    root = tk.Tk()
    #root.overrideredirect(True)
    root.attributes('-topmost', 1)  # 窗口置于顶层
    root.title('计时器v1.31 作者圈名：七号')
    timers = []
    # 根据 Excel 数据创建计时器

    index_timer = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if(row[0] == ""):
            continue
        hotkeys = str(row[2]).split(' ')  # 假设快捷键在第二列（索引为1）
        for index, hotkey in enumerate(hotkeys):
            if hotkey in num_to_symbol:
                hotkeys[index] = num_to_symbol[hotkey]
        
        timer = CountdownTimer(root, str(row[0]), int(row[1]), 0, index_timer, hotkeys, int(row[3]), int(row[4]), int(row[5]))
        timer.remaining_time.set(f"{str(row[0])}: {str(row[2])}")
        timers.append(timer)
        index_timer+=1



    '''
    timer0 = CountdownTimer(root, "7 5 称 号", 60, 0, 0, '$')
    timer1 = CountdownTimer(root, "斗 志", 120, 0, 1, '%')
    timer2 = CountdownTimer(root, "5 6 称 号", 30, 0, 2, '&')
    timer3 = CountdownTimer(root, "超越栏", 20, 0, 3, '*')'''
    '''
    timer0_button = tk.Button(root, text="更改快捷键", command=timer0.change_hotkey_window)
    timer0_button.grid(row=4, column=0)

    timer1_button = tk.Button(root, text="更改快捷键", command=timer1.change_hotkey_window)
    timer1_button.grid(row=4, column=1)

    timer2_button = tk.Button(root, text="更改快捷键", command=timer2.change_hotkey_window)
    timer2_button.grid(row=4, column=2)
    
    timer3_button = tk.Button(root, text="更改快捷键", command=timer3.change_hotkey_window)
    timer3_button.grid(row=4, column=3)
    '''
    
    keyboard.on_press_key('-', on_test_press)
    #keyboard.on_press_key('F9', on_test_press)

    #keyboard.add_hotkey('alt+1', on_f2_press)
    #root.bind("<F2>", on_f2_press)
    if(font_size_flag == 0 and vertical_flag != 1):
        test_label = tk.Label(root, text="调整完字体大小后自己调整窗口大小")
        test_label.grid(row=3, column=0,columnspan=2)
        increase_button = tk.Button(root, text="增加字体大小", command=increase_font_size)
        increase_button.grid(row=3, column=2)
        decrease_button = tk.Button(root, text="减小字体大小", command=decrease_font_size)
        decrease_button.grid(row=3, column=3)
    #close_button = tk.Button(root, text="关闭", command=close_window)
    #close_button.grid(row=0, column=8, columnspan=6)  # 使用grid布局管理器
    
    root.mainloop()